"""Extract, normalize, group, and enrich network-site IDs from Excel files.

This portfolio-safe version contains no operational data or organization names.
Column positions and region labels can be adjusted for an authorized dataset.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


PROVINCE_COLUMN = 8
CHILD_SITE_COLUMN = 11
GATEWAY_COLUMN = 12
IPRAN_COLUMN = 13

REGIONS = {
    "Region A": ("region a",),
    "Region B": ("region b", "region c"),
}


def parse_sites(value: object) -> list[str]:
    """Split a cell containing comma-, slash-, or whitespace-separated IDs."""
    if value is None:
        return []

    text = str(value).strip()
    if not text:
        return []

    return [item for item in re.split(r"[,/\s]+", text) if item]


def classify_region(province: object) -> str | None:
    """Map a province/state cell to one of the configured portfolio regions."""
    value = str(province or "").strip().lower()
    for region, keywords in REGIONS.items():
        if any(keyword in value for keyword in keywords):
            return region
    return None


def find_customer_site_id(worksheet, row: int) -> str | None:
    """Find a short site-like identifier in the columns before Child Site."""
    for column in range(1, CHILD_SITE_COLUMN):
        value = worksheet.cell(row=row, column=column).value
        if value is None:
            continue
        candidate = str(value).strip()
        if candidate and len(candidate) <= 20 and re.fullmatch(r"[\w-]+", candidate):
            return candidate
    return None


def iter_data_rows(worksheet, province_column: int = PROVINCE_COLUMN):
    """Yield rows until five consecutive empty province cells are encountered."""
    empty_rows = 0
    row = 2

    while empty_rows < 5:
        province = worksheet.cell(row=row, column=province_column).value
        if province is None or not str(province).strip():
            empty_rows += 1
        else:
            empty_rows = 0
            yield row, province
        row += 1


def process_primary_workbook(path: Path) -> dict[str, set[str]]:
    """Read child, gateway, and IPRAN site IDs from the primary workbook."""
    workbook = load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")
    worksheet = workbook.active
    grouped = {region: set() for region in REGIONS}

    for row, province in iter_data_rows(worksheet):
        region = classify_region(province)
        if region is None:
            continue

        child_site = worksheet.cell(row=row, column=CHILD_SITE_COLUMN).value
        if not parse_sites(child_site):
            child_site = find_customer_site_id(worksheet, row)
            if child_site:
                worksheet.cell(row=row, column=CHILD_SITE_COLUMN).value = child_site

        if not parse_sites(child_site):
            continue

        for column in (CHILD_SITE_COLUMN, GATEWAY_COLUMN, IPRAN_COLUMN):
            grouped[region].update(parse_sites(worksheet.cell(row=row, column=column).value))

    workbook.save(path)
    return grouped


def process_secondary_workbook(path: Path) -> dict[str, set[str]]:
    """Read IPRAN site IDs from the secondary workbook."""
    workbook = load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")
    worksheet = workbook.active
    grouped = {region: set() for region in REGIONS}

    for row, province in iter_data_rows(worksheet):
        region = classify_region(province)
        if region:
            grouped[region].update(
                parse_sites(worksheet.cell(row=row, column=IPRAN_COLUMN).value)
            )

    return grouped


def load_site_mapping(path: Path) -> dict[str, str]:
    """Load a two-column Site ID -> Site Name mapping workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    mapping: dict[str, str] = {}

    for site_id, site_name, *_ in worksheet.iter_rows(min_row=2, values_only=True):
        if site_id is None:
            continue
        normalized_id = str(site_id).strip()
        mapping[normalized_id] = str(site_name or normalized_id).strip()

    return mapping


def save_region_report(
    sites: set[str], mapping: dict[str, str], destination: Path
) -> None:
    """Save a sorted and formatted site report."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sites"
    worksheet.append(["Site ID", "Site Name"])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for site_id in sorted(sites, key=str.casefold):
        worksheet.append([site_id, mapping.get(site_id, site_id)])

    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 45
    worksheet.freeze_panes = "A2"
    workbook.save(destination)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("primary", type=Path, help="Primary work-permit workbook")
    parser.add_argument("secondary", type=Path, help="Secondary work-permit workbook")
    parser.add_argument("mapping", type=Path, help="Site ID-to-name mapping workbook")
    parser.add_argument("--output", type=Path, default=Path("output"))
    args = parser.parse_args()

    args.output.mkdir(parents=True, exist_ok=True)
    primary = process_primary_workbook(args.primary)
    secondary = process_secondary_workbook(args.secondary)
    mapping = load_site_mapping(args.mapping)

    for region in REGIONS:
        combined = primary[region] | secondary[region]
        safe_name = region.lower().replace(" ", "_")
        output_path = args.output / f"{safe_name}_sites.xlsx"
        save_region_report(combined, mapping, output_path)
        print(f"{region}: wrote {len(combined)} unique sites to {output_path}")


if __name__ == "__main__":
    main()

